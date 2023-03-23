import warnings
import openpyxl
import json
from faker import Faker
from parceiro.funcoesParceiro import *
from produto.funcoesProduto import *

def cadastrar(projName, telaCadastro, tipoCadastro, argQuantidadeCadastro):
    checkTipoCadastro(checkProjName(projName), checkTelaCadastro(telaCadastro), tipoCadastro, argQuantidadeCadastro)

def checkTipoCadastro(projName, telaCadastro, tipoCadastro, argQuantidadeCadastro):
    if tipoCadastro == "-op":
        preencherPlanilhaTodosCampos(argQuantidadeCadastro, telaCadastro, loadCamposJson(projName, telaCadastro))
    if tipoCadastro == "-o":
        preencherPlanilhaCamposObrigatorios(argQuantidadeCadastro, telaCadastro, loadCamposJson(projName, telaCadastro))

def checkValorPreencher(infDadoCriado, campoAlgoritmo, campoCabecalho, telaCadastro):
    if telaCadastro == "parceiros":
        return valorParceiro(infDadoCriado, campoAlgoritmo, campoCabecalho)
    if telaCadastro == "produtos":
        return valorProduto(campoAlgoritmo)
    
def checkDataBase(telaCadastro, ambiente,  contmaticId):
    if telaCadastro == "-par":
        get_info_parceiros(telaCadastro, ambiente, contmaticId)
    if telaCadastro == "-pro":
        get_info_produtos(telaCadastro, ambiente, contmaticId)  
    
def loadPlanilha(telacadastro):
    warnings.simplefilter(action='ignore', category=UserWarning)
    planilha = openpyxl.load_workbook(f'excel/planilhas/importar-{telacadastro}.xlsx')
    paginaDaPlanilha = planilha[f'importar-{telacadastro}']
    return planilha, paginaDaPlanilha

def loadCamposJson(projName, telaCadastro):
    with open("script-py/mapeamentos-json/mapeamento_cadastros.json", encoding='utf-8') as meu_json:
        dadosJson = json.load(meu_json) 
    camposJson = dadosJson[f'{projName}'][f'{telaCadastro}']["campos"]
    return camposJson

def checkTelaCadastro(telacadastro):
    if telacadastro == "-pro":
        telacadastro = "produtos"
    if telacadastro == "-par":
        telacadastro = "parceiros"
    return telacadastro

def checkProjName(projName):
    if projName == "-s":
        projName = "simplifique"
    return projName

def preencherPlanilhaCamposObrigatorios(argQuantidadeCadastro, telaCadastro, camposJson):
    planilha, paginaDaPlanilha = loadPlanilha(telaCadastro)
    for linhaPlanilha in range (POSICAO_PRIMEIRA_LINHA_PREENCHIMENTO, calcQuantidadeCadastro(argQuantidadeCadastro) ):
        infDadoCriado = {}
        for colunaPlanilha, campo in [(y + 1, campo) for (y, campo) in enumerate(camposJson)]:
            if campo["obrigatorio"]:
                paginaDaPlanilha.cell(row = linhaPlanilha, column = colunaPlanilha , value = checkValorPreencher(infDadoCriado, campo['algoritmo'], campo['cabecalho'], telaCadastro))
            else:
                paginaDaPlanilha.cell(row = linhaPlanilha, column = colunaPlanilha, value = "")
    planilha.save(f'excel/planilhasGerada/importar-{telaCadastro}.xlsx')

def preencherPlanilhaTodosCampos(argQuantidadeCadastro, telaCadastro, camposJson): 
    planilha, paginaDaPlanilha = loadPlanilha(telaCadastro)
    for linhaPlanilha in range (POSICAO_PRIMEIRA_LINHA_PREENCHIMENTO, calcQuantidadeCadastro(argQuantidadeCadastro) ):
        infDadoCriado = {}
        for colunaPlanilha, campo in [(y + 1, campo) for (y, campo) in enumerate(camposJson)]:
            paginaDaPlanilha.cell(row = linhaPlanilha, column = colunaPlanilha , value = checkValorPreencher(infDadoCriado, campo['algoritmo'],  campo['cabecalho'], telaCadastro))
    planilha.save(f'excel/planilhasGerada/importar-{telaCadastro}.xlsx')