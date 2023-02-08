import openpyxl
import json
from faker import Faker
from funcoes import *

# main(projName, TelaCadastro, quantidadeCadastros)

planilha = openpyxl.load_workbook('excel/planilhas/importar-parceiro.xlsx')
paginaParceiro = planilha['importar-parceiros']


with open("script-py/mapeamento_cadastros.json", encoding='utf-8') as meu_json:
    dadosJson = json.load(meu_json)

todosCampos = dadosJson["simplifique"]["clientes"]["campos"]

def preencherPlanilhaCamposObrigatorios(argQuantidadeCadastro):
    for linhaPlanilha in range (POSICAO_PRIMEIRA_LINHA_PREENCHIMENTO, calcQuantidadeCadastro(argQuantidadeCadastro) ):
        infDadoCriado = {}
        for colunaPlanilha, campo in [(y + 1, campo) for (y, campo) in enumerate(todosCampos)]:
            if campo["obrigatorio"]:
                campoAlgoritmo = campo['algoritmo']
                campoCabecalho =  campo['cabecalho']
                paginaParceiro.cell(row = linhaPlanilha, column = colunaPlanilha , value = checkValor(infDadoCriado, campoAlgoritmo, campoCabecalho))

def preencherPlanilhaTodosCampos(argQuantidadeCadastro): 
    for linhaPlanilha in range (POSICAO_PRIMEIRA_LINHA_PREENCHIMENTO, calcQuantidadeCadastro(argQuantidadeCadastro) ):
        infDadoCriado = {}
        for colunaPlanilha, campo in [(y + 1, campo) for (y, campo) in enumerate(todosCampos)]:
            campoAlgoritmo = campo['algoritmo']
            campoCabecalho =  campo['cabecalho']
            paginaParceiro.cell(row = linhaPlanilha, column = colunaPlanilha , value = checkValor(infDadoCriado, campoAlgoritmo, campoCabecalho))

preencherPlanilhaTodosCampos(10)
planilha.save('excel/planilhasGerada/importar-parceiros-teste.xlsx')